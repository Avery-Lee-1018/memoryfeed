#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
EXPORTS_DIR = ROOT / "exports"
RAW_ALL_PATH = EXPORTS_DIR / "feed_audit_raw_all_items.json"
RAW_FEED_PATH = EXPORTS_DIR / "feed_audit_raw.json"
CSV_PATH = EXPORTS_DIR / "feed_audit_template.csv"
XLSX_PATH = EXPORTS_DIR / "feed_audit_template.xlsx"
README_PATH = EXPORTS_DIR / "feed_audit_readme.txt"


def load_rows(path: Path) -> list[dict[str, Any]]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(raw, list) and raw and isinstance(raw[0], dict):
        results = raw[0].get("results", [])
        if isinstance(results, list):
            return results
    return []


def n_int(v: Any) -> int:
    try:
        return int(v)
    except Exception:
        return 0


def n_str(v: Any) -> str:
    return "" if v is None else str(v)


def normalized_items() -> list[dict[str, Any]]:
    src_path = RAW_ALL_PATH if RAW_ALL_PATH.exists() else RAW_FEED_PATH
    rows = load_rows(src_path)
    normalized: list[dict[str, Any]] = []
    for r in rows:
        normalized.append(
            {
                "source_id": n_int(r.get("source_id")),
                "source_name": n_str(r.get("source_name")),
                "source_url": n_str(r.get("source_url")),
                "source_active": n_int(r.get("is_active", 1)),
                "source_level": n_str(r.get("level")),
                "source_added_at": n_str(r.get("source_added_at")),
                "item_id": n_int(r.get("item_id")),
                "item_url": n_str(r.get("item_url", r.get("url"))),
                "title": n_str(r.get("title")),
                "summary": n_str(r.get("summary")),
                "thumbnail_url": n_str(r.get("thumbnail_url")),
                "item_status": n_str(r.get("item_status", r.get("status"))),
                "exposure_count": n_int(r.get("exposure_count")),
                "first_exposed_date": n_str(r.get("first_exposed_date")),
                "last_exposed_date": n_str(r.get("last_exposed_date")),
                "seen_since_2026_04_01": n_int(r.get("seen_since_2026_04_01")),
                "has_note": n_int(r.get("has_note")),
                "note_updated_at": n_str(r.get("note_updated_at")),
                "split_items": n_int(r.get("split_items")),
                "extraction_mode": n_str(r.get("extraction_mode")),
                "landing_suspect": n_int(r.get("landing_suspect")),
                "thumb_suspect": n_int(r.get("thumb_suspect", r.get("weak_thumbnail"))),
                "title_suspect": n_int(r.get("title_suspect", r.get("weak_title"))),
                "summary_suspect": n_int(r.get("summary_suspect", r.get("weak_summary"))),
            }
        )
    normalized.sort(key=lambda x: (x["source_added_at"], x["source_name"].lower(), x["item_id"]), reverse=True)
    return normalized


def source_summary(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_source: dict[int, dict[str, Any]] = {}
    for item in items:
        sid = item["source_id"]
        if sid not in by_source:
            by_source[sid] = {
                "source_id": sid,
                "source_name": item["source_name"],
                "source_url": item["source_url"],
                "source_active": item["source_active"],
                "source_level": item["source_level"],
                "source_added_at": item["source_added_at"],
                "item_count": 0,
                "exposed_item_count": 0,
                "memo_item_count": 0,
                "latest_exposed_date": "",
            }
        s = by_source[sid]
        s["item_count"] += 1
        if item["exposure_count"] > 0:
            s["exposed_item_count"] += 1
        if item["has_note"] == 1:
            s["memo_item_count"] += 1
        if item["last_exposed_date"] and item["last_exposed_date"] > s["latest_exposed_date"]:
            s["latest_exposed_date"] = item["last_exposed_date"]
    return sorted(by_source.values(), key=lambda x: (x["source_added_at"], x["source_name"].lower()), reverse=True)


def write_csv(items: list[dict[str, Any]]) -> None:
    headers = [
        "review_status",
        "source_review_status",
        "source_id",
        "source_name",
        "source_url",
        "source_check",
        "item_id",
        "item_url",
        "landing_check",
        "title",
        "title_check",
        "summary",
        "description_check",
        "thumbnail_url",
        "thumbnail_check",
        "issue_category",
        "fix_priority",
        "issue_notes",
        "source_active",
        "source_level",
        "source_added_at",
        "item_status",
        "exposure_count",
        "first_exposed_date",
        "last_exposed_date",
        "seen_since_2026_04_01",
        "has_note",
        "note_updated_at",
        "split_items",
        "extraction_mode",
        "landing_suspect",
        "thumb_suspect",
        "title_suspect",
        "summary_suspect",
    ]
    with CSV_PATH.open("w", encoding="utf-8-sig", newline="") as fp:
        w = csv.writer(fp)
        w.writerow(headers)
        for i in items:
            w.writerow(
                [
                    "미확인",
                    "",
                    i["source_id"],
                    i["source_name"],
                    i["source_url"],
                    "",
                    i["item_id"],
                    i["item_url"],
                    "",
                    i["title"],
                    "",
                    i["summary"],
                    "",
                    i["thumbnail_url"],
                    "",
                    "",
                    "",
                    "",
                    i["source_active"],
                    i["source_level"],
                    i["source_added_at"],
                    i["item_status"],
                    i["exposure_count"],
                    i["first_exposed_date"],
                    i["last_exposed_date"],
                    i["seen_since_2026_04_01"],
                    i["has_note"],
                    i["note_updated_at"],
                    i["split_items"],
                    i["extraction_mode"],
                    i["landing_suspect"],
                    i["thumb_suspect"],
                    i["title_suspect"],
                    i["summary_suspect"],
                ]
            )


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="111827")
    for c in ws[1]:
        c.font = Font(color="FFFFFF", bold=True)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(1, ws.max_column).column_letter}{ws.max_row}"


def add_dv(ws, cell_range: str, formula: str) -> None:
    dv = DataValidation(type="list", formula1=f"={formula}", allow_blank=True)
    dv.error = "목록에서 선택해주세요."
    ws.add_data_validation(dv)
    dv.add(cell_range)


def write_xlsx(items: list[dict[str, Any]]) -> None:
    wb = Workbook()

    guide = wb.active
    guide.title = "Guide"
    guide.append(["단계", "가이드"])
    guide.append(["1", "Source Check에서 소스 단위로 상태를 체크하세요."])
    guide.append(["2", "Item Audit에서 각 항목 바로 옆 체크칸(landing/title/description/thumbnail)을 채우세요."])
    guide.append(["3", "source_review_status는 Source Check 상태를 자동 참조합니다."])
    guide.column_dimensions["A"].width = 12
    guide.column_dimensions["B"].width = 100

    src = wb.create_sheet("Source Check")
    src_headers = [
        "review_status",
        "source_id",
        "source_name",
        "source_url",
        "source_check",
        "landing_check",
        "thumbnail_check",
        "title_check",
        "description_check",
        "source_issue_notes",
        "source_active",
        "source_level",
        "source_added_at",
        "item_count",
        "exposed_item_count",
        "memo_item_count",
        "latest_exposed_date",
    ]
    src.append(src_headers)
    for s in source_summary(items):
        src.append(
            [
                "미확인",
                s["source_id"],
                s["source_name"],
                s["source_url"],
                "",
                "",
                "",
                "",
                "",
                "",
                s["source_active"],
                s["source_level"],
                s["source_added_at"],
                s["item_count"],
                s["exposed_item_count"],
                s["memo_item_count"],
                s["latest_exposed_date"],
            ]
        )
    style_header(src)
    src_widths = {
        "A": 12,
        "B": 9,
        "C": 20,
        "D": 42,
        "E": 12,
        "F": 12,
        "G": 12,
        "H": 12,
        "I": 14,
        "J": 36,
        "K": 10,
        "L": 10,
        "M": 20,
        "N": 10,
        "O": 12,
        "P": 11,
        "Q": 14,
    }
    for c, w in src_widths.items():
        src.column_dimensions[c].width = w

    item = wb.create_sheet("Item Audit")
    item_headers = [
        "review_status",
        "source_review_status",
        "source_id",
        "source_name",
        "source_url",
        "source_check",
        "item_id",
        "item_url",
        "landing_check",
        "title",
        "title_check",
        "summary",
        "description_check",
        "thumbnail_url",
        "thumbnail_check",
        "issue_category",
        "fix_priority",
        "issue_notes",
        "source_active",
        "source_level",
        "source_added_at",
        "item_status",
        "exposure_count",
        "first_exposed_date",
        "last_exposed_date",
        "seen_since_2026_04_01",
        "has_note",
        "note_updated_at",
        "split_items",
        "extraction_mode",
        "landing_suspect",
        "thumb_suspect",
        "title_suspect",
        "summary_suspect",
    ]
    item.append(item_headers)
    for idx, r in enumerate(items, start=2):
        item.append(
            [
                "미확인",
                f'=IFERROR(INDEX(\'Source Check\'!$A:$A, MATCH(C{idx}, \'Source Check\'!$B:$B, 0)), "미등록")',
                r["source_id"],
                r["source_name"],
                r["source_url"],
                f'=IFERROR(INDEX(\'Source Check\'!$E:$E, MATCH(C{idx}, \'Source Check\'!$B:$B, 0)), "")',
                r["item_id"],
                r["item_url"],
                "",
                r["title"],
                "",
                r["summary"],
                "",
                r["thumbnail_url"],
                "",
                "",
                "",
                "",
                r["source_active"],
                r["source_level"],
                r["source_added_at"],
                r["item_status"],
                r["exposure_count"],
                r["first_exposed_date"],
                r["last_exposed_date"],
                r["seen_since_2026_04_01"],
                r["has_note"],
                r["note_updated_at"],
                r["split_items"],
                r["extraction_mode"],
                r["landing_suspect"],
                r["thumb_suspect"],
                r["title_suspect"],
                r["summary_suspect"],
            ]
        )
    style_header(item)
    item.freeze_panes = "H2"
    item_widths = {
        "A": 12,
        "B": 14,
        "C": 9,
        "D": 20,
        "E": 34,
        "F": 12,
        "G": 9,
        "H": 42,
        "I": 12,
        "J": 30,
        "K": 12,
        "L": 44,
        "M": 14,
        "N": 36,
        "O": 12,
        "P": 14,
        "Q": 10,
        "R": 40,
        "S": 10,
        "T": 10,
        "U": 20,
        "V": 10,
        "W": 10,
        "X": 12,
        "Y": 12,
        "Z": 10,
        "AA": 8,
        "AB": 18,
        "AC": 10,
        "AD": 13,
        "AE": 10,
        "AF": 10,
        "AG": 10,
    }
    for c, w in item_widths.items():
        item.column_dimensions[c].width = w

    subtle_fill = PatternFill("solid", fgColor="F8FAFC")
    for row in item.iter_rows(min_row=2, max_row=item.max_row, min_col=19, max_col=34):
        for cell in row:
            cell.fill = subtle_fill

    opt = wb.create_sheet("Options")
    values = {
        "A": ["미확인", "확인완료", "이슈있음", "보류"],
        "B": ["정상", "미작동", "보류"],
        "C": ["정상", "404", "리다이렉트 오류", "접근 차단", "기타"],
        "D": ["정상", "기본 이미지", "깨짐", "없음", "기타"],
        "E": ["정상", "URL 노출", "깨짐/인코딩", "무의미 텍스트", "없음"],
        "F": ["정상", "특수문자 노이즈", "요약 없음", "본문 오염", "없음"],
        "G": ["landing", "thumbnail", "title", "description", "dedupe", "split", "performance", "other"],
        "H": ["P0", "P1", "P2", "P3"],
    }
    for col, arr in values.items():
        for idx, v in enumerate(arr, start=1):
            opt[f"{col}{idx}"] = v
    opt.sheet_state = "hidden"

    src_max = src.max_row
    item_max = item.max_row
    add_dv(src, f"A2:A{src_max}", "Options!$A$1:$A$4")
    add_dv(src, f"E2:E{src_max}", "Options!$B$1:$B$3")
    add_dv(src, f"F2:F{src_max}", "Options!$C$1:$C$5")
    add_dv(src, f"G2:G{src_max}", "Options!$D$1:$D$5")
    add_dv(src, f"H2:H{src_max}", "Options!$E$1:$E$5")
    add_dv(src, f"I2:I{src_max}", "Options!$F$1:$F$5")

    add_dv(item, f"A2:A{item_max}", "Options!$A$1:$A$4")
    add_dv(item, f"F2:F{item_max}", "Options!$B$1:$B$3")
    add_dv(item, f"I2:I{item_max}", "Options!$C$1:$C$5")
    add_dv(item, f"K2:K{item_max}", "Options!$E$1:$E$5")
    add_dv(item, f"M2:M{item_max}", "Options!$F$1:$F$5")
    add_dv(item, f"O2:O{item_max}", "Options!$D$1:$D$5")
    add_dv(item, f"P2:P{item_max}", "Options!$G$1:$G$8")
    add_dv(item, f"Q2:Q{item_max}", "Options!$H$1:$H$4")

    wb.save(XLSX_PATH)


def main() -> None:
    items = normalized_items()
    if not items:
        raise RuntimeError("No rows loaded from raw export json.")
    write_csv(items)
    write_xlsx(items)
    source_count = len({i["source_id"] for i in items})
    README_PATH.write_text(
        "\n".join(
            [
                "Feed audit export generated",
                f"- generated_at: {datetime.now().isoformat()}",
                f"- source_json: {RAW_ALL_PATH if RAW_ALL_PATH.exists() else RAW_FEED_PATH}",
                f"- total_items: {len(items)}",
                f"- total_sources: {source_count}",
                f"- files:",
                f"  - {CSV_PATH}",
                f"  - {XLSX_PATH}",
                "- sheets:",
                "  - Guide",
                "  - Source Check",
                "  - Item Audit",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    print(f"items={len(items)}")
    print(f"sources={source_count}")
    print(CSV_PATH)
    print(XLSX_PATH)


if __name__ == "__main__":
    main()
